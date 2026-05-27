"""
Extração de texto plano a partir dos documentos do inventário.

Funciona como um **dispatcher por extensão de arquivo**: a função pública
``extract_from_scanned_file`` identifica o tipo do arquivo e delega para o
extrator específico. Todos os erros de parsing são capturados internamente e
devolvidos em ``ExtractOutcome.ok = False``, para que uma falha em um arquivo
não interrompa o pipeline inteiro de indexação.

Formatos suportados
-------------------
- ``.docx`` — parágrafos + tabelas em ordem de documento (python-docx)
- ``.xlsx`` / ``.xlsm`` — todas as abas (openpyxl, modo read-only)
- ``.pdf`` — texto por página (pypdf)
- ``.txt`` / ``.md`` — texto plano com detecção de encoding
- ``.csv`` — células separadas por tabulação

Truncagem
---------
Todos os extratores respeitam ``max_chars_total`` para evitar carregar arquivos
muito grandes inteiramente na RAM. Arquivos truncados são indexados normalmente
(o texto disponível já cobre a maioria dos casos práticos).
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from projects_loader import ScannedFile


# ── Resultado da extração ────────────────────────────────────────────────────

@dataclass
class ExtractOutcome:
    """
    Resultado de uma tentativa de extração de texto.

    Atributos
    ---------
    text:
        Texto bruto extraído. Vazio se a extração falhou ou o arquivo não
        contém texto legível.
    detail:
        Mensagem humana descrevendo o resultado: número de parágrafos/tabelas,
        abas processadas, aviso de truncagem ou mensagem de erro. Exibida na
        UI de indexação e gravada nos metadados do chunk.
    ok:
        ``True`` se o texto foi extraído com sucesso (mesmo que truncado).
        ``False`` indica falha de parsing ou arquivo sem conteúdo.
    """

    text: str
    detail: str
    ok: bool


# ── Roteador principal ───────────────────────────────────────────────────────

def extract_from_scanned_file(sf: ScannedFile, *, max_chars_total: int = 2_000_000) -> ExtractOutcome:
    """
    Seleciona o extrator adequado para ``sf`` e retorna o texto plano.

    Qualquer exceção de parsing é capturada e devolvida como
    ``ExtractOutcome(ok=False)``, para que o pipeline de indexação continue
    processando os demais arquivos mesmo quando um documento está corrompido
    ou tem formato inesperado.

    Parâmetros
    ----------
    sf:
        Arquivo escaneado com caminho absoluto e metadados.
    max_chars_total:
        Limite de caracteres do texto final. Evita alocar GBs de memória para
        PDFs ou planilhas muito grandes. O padrão de 2 M chars cobre a grande
        maioria dos documentos de laboratório.
    """
    path = sf.absolute_path
    suffix = path.suffix.lower()

    try:
        if suffix == ".docx":
            return _extract_docx(path, max_chars_total=max_chars_total)
        if suffix in (".xlsx", ".xlsm"):
            return _extract_excel(path, max_chars_total=max_chars_total)
        if suffix == ".pdf":
            return _extract_pdf(path, max_chars_total=max_chars_total)
        if suffix in (".txt", ".md"):
            return _extract_plain(path, max_chars_total=max_chars_total)
        if suffix == ".csv":
            return _extract_csv(path, max_chars_total=max_chars_total)
    except Exception as exc:  # noqa: BLE001 — falha isolada; pipeline não deve parar
        return ExtractOutcome(text="", detail=f"Erro ao ler arquivo: {exc}", ok=False)

    return ExtractOutcome(text="", detail=f"Extensão não suportada para extração: {suffix}", ok=False)


# ── Helper de truncagem ──────────────────────────────────────────────────────

def _truncate(s: str, limit: int) -> tuple[str, bool]:
    """Retorna ``(texto, foi_truncado)``."""
    if len(s) <= limit:
        return s, False
    return s[:limit], True


# ── Helpers para documentos Word (.docx) ────────────────────────────────────

def _walk_block_container(element: object, parent: object):
    """
    Percorre parágrafos e tabelas em ordem de documento dentro de um contêiner.

    Além de ``<w:p>`` e ``<w:tbl>`` diretos, desce em:
    - ``<w:sdt>`` (controles de conteúdo / formulários Word — comum em templates)
    - ``<w:ins>`` (revisões rastreadas)
    - ``<w:smartTag>`` (marcadores legados)

    Sem isso, tabelas de insumos embutidas em controles de conteúdo ficam invisíveis
    para a busca semântica mesmo com ``python-docx`` instalado.
    """
    from docx.oxml.ns import qn
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    for child in element.iterchildren():  # type: ignore[attr-defined]
        tag = child.tag
        if tag == qn("w:p"):
            yield Paragraph(child, parent)
        elif tag == qn("w:tbl"):
            yield Table(child, parent)
        elif tag == qn("w:sdt"):
            content = child.find(qn("w:sdtContent"))
            if content is not None:
                yield from _walk_block_container(content, parent)
        elif tag in (qn("w:ins"), qn("w:smartTag")):
            yield from _walk_block_container(child, parent)


def _iter_docx_blocks(document: object):
    """Blocos do corpo principal do documento (``<w:body>``)."""
    yield from _walk_block_container(document.element.body, document)  # type: ignore[attr-defined]


def _iter_hdr_ftr_blocks(header_footer: object):
    """Blocos de cabeçalho ou rodapé (quando não vinculado ao anterior)."""
    if getattr(header_footer, "is_linked_to_previous", False):
        return
    yield from _walk_block_container(header_footer._element, header_footer)  # type: ignore[attr-defined]


def _normalize_cell_text(text: str) -> str:
    """
    Normaliza o texto de uma célula para uso em pareamentos ``coluna: valor``.

    Operações
    ---------
    1. Quebras de linha internas (``\\n``/``\\r``) viram ``"; "`` para a
       célula caber em uma única linha do extrato — preservando o pareamento
       coluna-valor mesmo quando a original tinha múltiplos itens
       (ex.: ``"EQA0506; EQA0982"`` em colunas de equipamento).
    2. Tokens duplicados consecutivos (após ``split(";")``) são deduplicados.
       Isso elimina ruído típico de células ancorando ``<w:vMerge>``, onde
       ``python-docx`` repete o conteúdo da âncora para cada linha mesclada
       (ex.: ``"NA; NA; NA"`` vira ``"NA"``).
    3. Separadores vazios (``";; "``) são colapsados em ``"; "``.
    """
    if not text:
        return ""
    cleaned = text.replace("\r\n", "\n").replace("\r", "\n")
    parts = [p.strip().rstrip(";").strip() for p in cleaned.split("\n")]
    parts = [p for p in parts if p]

    deduped: list[str] = []
    for p in parts:
        if not deduped or deduped[-1] != p:
            deduped.append(p)

    # Se todos os tokens forem idênticos (caso da vMerge âncora), mantém só um.
    if deduped and all(p == deduped[0] for p in deduped):
        deduped = [deduped[0]]

    return "; ".join(deduped)


def _cell_to_text(cell: object) -> str:
    """
    Texto de uma célula, incluindo parágrafos e tabelas aninhadas.

    ``cell.text`` do python-docx omite tabelas aninhadas; percorrer os blocos
    da célula garante que sub-tabelas de insumos entrem na indexação. O
    resultado passa por ``_normalize_cell_text`` para colapsar quebras de
    linha internas em ``"; "``, evitando "linhas-fantasma" no extrato final.
    """
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    pieces: list[str] = []
    for block in _walk_block_container(cell._tc, cell):  # type: ignore[attr-defined]
        if isinstance(block, Paragraph):
            t = (block.text or "").strip()
            if t:
                pieces.append(t)
        elif isinstance(block, Table):
            for line in _table_to_lines(block):
                pieces.append(line.replace("\t", " | "))
    if not pieces:
        t = (cell.text or "").strip()  # type: ignore[attr-defined]
        if t:
            pieces.append(t)
    return _normalize_cell_text(" | ".join(pieces))


def _cell_vmerge_state(cell: object) -> str | None:
    """
    Lê ``<w:vMerge>`` de uma célula e devolve seu estado.

    Retornos
    --------
    ``"restart"`` — célula é a âncora de uma mescla vertical (tem conteúdo).
    ``"continue"`` — célula é continuação de uma mescla vertical iniciada
        em linha anterior; nesse caso ``cell.text`` do python-docx repete o
        conteúdo da âncora, o que gera "linhas-fantasma" no extrato se a
        linha inteira for de continuação. Detectar permite pular a linha.
    ``None`` — célula sem mescla vertical.
    """
    from docx.oxml.ns import qn

    tcPr = cell._tc.find(qn("w:tcPr"))  # type: ignore[attr-defined]
    if tcPr is None:
        return None
    vMerge = tcPr.find(qn("w:vMerge"))
    if vMerge is None:
        return None
    val = vMerge.get(qn("w:val"))
    return "restart" if val == "restart" else "continue"


def _row_is_all_vmerge_continue(row: object) -> bool:
    """
    Indica se a linha é composta SÓ por células de continuação vertical.

    Linhas assim não trazem informação nova (são continuação visual da
    linha-âncora acima) e, se incluídas no extrato, viram pareamentos
    repetidos ou linhas com cara de ``"NA NA NA"`` que confundem o LLM.
    """
    has_cell = False
    for cell in row.cells:  # type: ignore[attr-defined]
        has_cell = True
        if _cell_vmerge_state(cell) != "continue":
            return False
    return has_cell


def _table_row_cell_texts(row: object) -> list[str]:
    """
    Retorna os textos das células de uma linha, sem duplicar células mescladas.

    Em tabelas Word com células mescladas (merge), ``row.cells`` pode retornar
    o mesmo objeto ``_tc`` (nó XML da célula) mais de uma vez. A deduplicação
    por ``id(cell._tc)`` garante que cada célula física apareça apenas uma vez
    na linha extraída.
    """
    seen_tc: set[int] = set()
    texts: list[str] = []
    for cell in row.cells:  # type: ignore[attr-defined]
        tc_id = id(cell._tc)
        if tc_id in seen_tc:
            continue
        seen_tc.add(tc_id)
        texts.append(_cell_to_text(cell))
    return texts


def _table_to_lines(table: object) -> list[str]:
    """Serializa uma tabela Word em linhas TSV (fallback legado).

    Mantida para serialização de tabelas aninhadas dentro de células (onde o
    formato ``coluna: valor`` por reagente seria confuso). O caminho principal
    do extrator passa por :func:`_table_to_records` /
    :func:`_serialize_table_as_records`.
    """
    lines: list[str] = []
    for row in table.rows:  # type: ignore[attr-defined]
        if _row_is_all_vmerge_continue(row):
            continue
        cells = _table_row_cell_texts(row)
        line = "\t".join(cells).strip()
        if line:
            lines.append(line)
    return lines


def _table_to_records(table: object) -> list[dict[str, str]]:
    """
    Detecta o cabeçalho (1ª linha) e retorna ``[{coluna: valor, ...}]`` por linha.

    A 1ª linha da tabela é tratada como header. Linhas de continuação vertical
    (``<w:vMerge>`` sem ``val="restart"``) são puladas — elas geram entradas
    redundantes (cell.text repete o conteúdo da linha-âncora) e poluem o
    extrato.

    Quando a tabela tem menos de 2 linhas ou o header é totalmente vazio,
    devolve lista vazia; o chamador faz fallback para o formato TSV legado.
    """
    rows = list(table.rows)  # type: ignore[attr-defined]
    if len(rows) < 2:
        return []

    header_cells = _table_row_cell_texts(rows[0])
    if not any(c.strip() for c in header_cells):
        return []

    headers = [c.strip() or f"Coluna{i + 1}" for i, c in enumerate(header_cells)]

    records: list[dict[str, str]] = []
    for row in rows[1:]:
        if _row_is_all_vmerge_continue(row):
            continue
        cells = _table_row_cell_texts(row)
        if not any(c.strip() for c in cells):
            continue
        # Normaliza tamanho: completa com vazio ou trunca à largura do header
        if len(cells) < len(headers):
            cells = cells + [""] * (len(headers) - len(cells))
        elif len(cells) > len(headers):
            cells = cells[: len(headers)]
        records.append({h: cells[i].strip() for i, h in enumerate(headers)})

    return records


def _serialize_table_as_records(
    table: object,
    *,
    title: str,
) -> list[str]:
    """
    Serializa cada linha da tabela como um item auto-contido ``coluna: valor``.

    Por que esse formato?
    ---------------------
    O extrator antigo emitia uma linha TSV por reagente, com o header da
    tabela só na 1ª linha. O ``chunk_text`` (limite ~520 chars) cortava
    tabelas longas no meio, e os chunks subsequentes ficavam sem o header
    — o LLM precisava adivinhar qual coluna era qual. Isso causava respostas
    erradas em perguntas do tipo "qual a validade do X?".

    Aqui, cada linha do extrato carrega TODOS os pareamentos
    (``Reagente: X · Fabricante/Código: Y · Lote/ativo: Z · Validade: W``).
    Mesmo que o RAG recupere um chunk com apenas 1 item, o pareamento
    coluna-valor está completo — não é necessário "lembrar" o header.

    Cai no formato TSV de ``_table_to_lines`` se a tabela não tiver header
    detectável (tabela de 1 linha, layout só visual, etc.).
    """
    records = _table_to_records(table)
    if not records:
        # Fallback: tabela sem header claro → mantém TSV puro para não perder dados.
        legacy = _table_to_lines(table)
        if not legacy:
            return []
        return [f"### {title}", *legacy]

    columns = list(records[0].keys())
    lines: list[str] = [f"### {title}", "Colunas: " + " · ".join(columns)]
    for i, rec in enumerate(records, start=1):
        parts: list[str] = []
        for col in columns:
            val = (rec.get(col) or "").strip()
            parts.append(f"{col}: {val if val else '(em branco)'}")
        lines.append(f"Item {i} — " + " · ".join(parts))
    return lines


def _infer_table_title(prev_paragraph: str) -> str:
    """
    Usa o último parágrafo não vazio antes da tabela como título, quando ele
    parece um cabeçalho de seção.

    Heurística conservadora: aceita parágrafos curtos (< 120 chars) e sem
    pontuação pesada (no máximo 1 ponto final, 2 vírgulas). Caso contrário,
    devolve o rótulo genérico ``"Tabela de insumos / materiais"`` — mantém
    compatibilidade com a sinalização já reconhecida pelo LLM.
    """
    fallback = "Tabela de insumos / materiais"
    t = (prev_paragraph or "").strip().rstrip(":").strip()
    if not t or len(t) > 120:
        return fallback
    if t.count(".") > 1 or t.count(",") > 2:
        return fallback
    return f"Tabela — {t}"


def _append_docx_blocks(
    blocks: object,
    parts: list[str],
    *,
    n_paras: list[int],
    n_tables: list[int],
) -> None:
    """
    Acumula parágrafos e tabelas serializadas em ``parts``.

    Para cada tabela, usa o último parágrafo não-vazio anterior como pista
    para um título descritivo (``_infer_table_title``). O conteúdo da tabela
    é serializado no formato ``Item N — coluna: valor · ...`` por linha de
    reagente, garantindo pareamento auto-contido (ver
    :func:`_serialize_table_as_records`).
    """
    from docx.table import Table
    from docx.text.paragraph import Paragraph

    last_paragraph_text = ""
    for block in blocks:
        if isinstance(block, Paragraph):
            t = (block.text or "").strip()
            if t:
                parts.append(t)
                n_paras[0] += 1
                last_paragraph_text = t
        elif isinstance(block, Table):
            title = _infer_table_title(last_paragraph_text)
            table_lines = _serialize_table_as_records(block, title=title)
            if table_lines:
                parts.extend(table_lines)
                n_tables[0] += 1
            last_paragraph_text = ""


# ── Extratores por formato ───────────────────────────────────────────────────

def _extract_docx(path: Path, *, max_chars_total: int) -> ExtractOutcome:
    """
    Extrai texto de um arquivo Word, preservando parágrafos e tabelas em ordem.

    Cada tabela recebe um cabeçalho ``### Tabela de insumos / materiais`` para
    sinalizar ao modelo de linguagem que o bloco seguinte é tabular — útil para
    perguntas sobre lotes, validades e fabricantes.
    """
    from docx import Document

    doc = Document(str(path))
    parts: list[str] = []
    n_tables = [0]
    n_paras = [0]

    _append_docx_blocks(_iter_docx_blocks(doc), parts, n_paras=n_paras, n_tables=n_tables)

    for section in doc.sections:
        for hf in (section.header, section.footer):
            _append_docx_blocks(
                _iter_hdr_ftr_blocks(hf),
                parts,
                n_paras=n_paras,
                n_tables=n_tables,
            )

    text = "\n".join(parts)
    n_paras_val, n_tables_val = n_paras[0], n_tables[0]
    text, cut = _truncate(text, max_chars_total)
    detail = f"docx: {n_paras_val} parágrafo(s), {n_tables_val} tabela(s)"
    if cut:
        detail += f" (truncado a {max_chars_total} caracteres)"
    return ExtractOutcome(text=text, detail=detail, ok=bool(text))


def _extract_excel(path: Path, *, max_chars_total: int) -> ExtractOutcome:
    """
    Extrai todas as abas de uma planilha Excel como texto tabular.

    Usa dois mecanismos de truncagem:
    1. Por linha: cada aba tem limite de 50 000 linhas para evitar loop muito
       longo em planilhas com muitas linhas em branco.
    2. Global: contador ``remaining`` para o total de caracteres. Quando
       esgotado, retorna imediatamente sem processar o restante das abas.

    Células individuais com mais de 2 000 chars são truncadas com ``…`` para
    evitar que fórmulas ou textos longos dominem o espaço de um único chunk.
    """
    from openpyxl import load_workbook

    # read_only=True e data_only=True: mais rápido e evita calcular fórmulas.
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    lines: list[str] = []
    remaining = max_chars_total

    try:
        for sheet in wb.worksheets:
            lines.append(f"### Planilha: {sheet.title}")
            row_limit = 50_000
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i > row_limit:
                    lines.append(f"(… planilha {sheet.title} truncada após {row_limit} linhas …)")
                    break
                cells = []
                for cell in row:
                    if cell is None:
                        cells.append("")
                    else:
                        s = str(cell).strip()
                        if len(s) > 2000:
                            s = s[:2000] + "…"
                        cells.append(s)
                line = "\t".join(cells).strip()
                if line:
                    lines.append(line)
                    remaining -= len(line) + 1
                    if remaining <= 0:
                        lines.append("(… truncado por limite global de caracteres …)")
                        text = "\n".join(lines)
                        return ExtractOutcome(
                            text=text,
                            detail="xlsx: múltiplas abas (truncado)",
                            ok=True,
                        )
    finally:
        wb.close()

    text = "\n".join(lines).strip()
    text, extra_cut = _truncate(text, max_chars_total)
    detail = "xlsx: abas e linhas"
    if extra_cut:
        detail += f" (truncado a {max_chars_total} caracteres)"
    if not text:
        return ExtractOutcome(text="", detail="xlsx: sem células com conteúdo", ok=False)
    return ExtractOutcome(text=text, detail=detail, ok=True)


def _extract_pdf(path: Path, *, max_chars_total: int) -> ExtractOutcome:
    """
    Extrai texto de um PDF página a página.

    Erros em páginas individuais (PDF corrompido, página de imagem sem OCR) são
    registrados como marcadores de texto ``[página N: erro ...]`` e o
    processamento continua nas demais páginas.
    """
    from pypdf import PdfReader

    reader = PdfReader(str(path))
    parts: list[str] = []
    for i, page in enumerate(reader.pages):
        try:
            t = page.extract_text() or ""
        except Exception as exc:  # noqa: BLE001
            parts.append(f"[página {i + 1}: erro {exc}]")
            continue
        if t.strip():
            parts.append(t)
    text = "\n\n".join(parts)
    text, cut = _truncate(text, max_chars_total)
    detail = "pdf: texto extraído"
    if cut:
        detail += f" (truncado a {max_chars_total} caracteres)"
    return ExtractOutcome(text=text, detail=detail, ok=bool(text))


def _extract_plain(path: Path, *, max_chars_total: int) -> ExtractOutcome:
    """
    Lê um arquivo de texto simples (.txt ou .md) com detecção de encoding.

    A ordem de tentativa é: UTF-8 → UTF-8 com BOM → latin-1. Latin-1 é o
    fallback final porque mapeia todos os 256 valores de byte sem lançar
    ``UnicodeDecodeError``, garantindo que o loop sempre termine com ``break``
    e que ``text`` sempre tenha um valor atribuído.
    """
    raw = path.read_bytes()
    text = ""
    for enc in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    text = text.strip()
    text, cut = _truncate(text, max_chars_total)
    detail = "texto plano"
    if cut:
        detail += f" (truncado a {max_chars_total} caracteres)"
    return ExtractOutcome(text=text, detail=detail, ok=bool(text))


def _extract_csv(path: Path, *, max_chars_total: int) -> ExtractOutcome:
    """
    Lê um CSV e serializa cada linha como campos separados por tabulação.

    Limite de 100 000 linhas por arquivo para evitar bloqueio prolongado em
    dumps de banco exportados acidentalmente.
    """
    import csv

    lines: list[str] = []
    remaining = max_chars_total
    with path.open("r", encoding="utf-8", errors="replace", newline="") as fh:
        reader = csv.reader(fh)
        for i, row in enumerate(reader):
            if i > 100_000:
                lines.append("(… CSV truncado após 100k linhas …)")
                break
            line = "\t".join(str(c) for c in row)
            lines.append(line)
            remaining -= len(line) + 1
            if remaining <= 0:
                lines.append("(… truncado por limite global …)")
                break
    text = "\n".join(lines).strip()
    text, cut = _truncate(text, max_chars_total)
    detail = "csv"
    if cut:
        detail += f" (truncado a {max_chars_total} caracteres)"
    return ExtractOutcome(text=text, detail=detail, ok=bool(text))
